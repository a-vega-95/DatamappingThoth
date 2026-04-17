"""
analyzer.py — Mapeador Inteligente de Archivos de Datos
Macro-Módulo 1-B: Analiza estructura (Names, Dtypes, Head, Tail)
de archivos CSV, Excel (XLSX/XLS/XLSM) y Parquet.
Genera informes en TXT (por defecto) o PDF.
"""

import os
import re
from typing import Dict, List, Optional, Callable, Tuple

from fpdf import FPDF

from src.config.settings import (
    EXTENSIONES_DATOS, CHUNK_SIZE, MUESTRA_VALORES,
    MAX_FILAS_BUSQUEDA_HEADER, HEAD_TAIL_ROWS,
)
from src.utils.file_utils import detectar_tipo_str, detectar_tipo_valor, formato_bytes
from src.utils.pdf_utils import compress_pdf


# ─────────────────────────────────────────────
# PDF interno para informes de datos
# ─────────────────────────────────────────────

class _PDFDatos(FPDF):
    """Reporte PDF para el mapeador de archivos de datos."""

    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'Mapa de Archivos de Datos — DatamappingThoth Pro', 0, 1, 'C')
        self.set_font('Arial', 'I', 8)
        self.cell(0, 5, 'Detección inteligente de encabezados | Head & Tail incluidos', 0, 1, 'C')
        self.ln(2)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')

    def archivo_header(self, nombre: str, tipo: str, ruta: str):
        self.set_fill_color(200, 220, 255)
        self.set_font('Arial', 'B', 10)
        self.cell(0, 7, f'ARCHIVO: {_clean(nombre)}', 0, 1, 'L', True)
        self.set_font('Arial', '', 8)
        self.cell(0, 5, f'Tipo: {tipo}', 0, 1)
        self.cell(0, 5, f'Ruta: {_clean(ruta)[:110]}', 0, 1)

    def hoja_header(self, nombre: str):
        self.set_fill_color(230, 230, 230)
        self.set_font('Arial', 'B', 9)
        self.cell(0, 6, f'  HOJA: {_clean(nombre)}', 0, 1, 'L', True)

    def info(self, texto: str):
        self.set_font('Arial', '', 8)
        self.cell(0, 5, _clean(texto)[:125], 0, 1)

    def error_msg(self, mensaje: str):
        self.set_font('Arial', 'I', 8)
        self.set_text_color(200, 0, 0)
        self.cell(0, 5, f'ERROR: {_clean(mensaje)[:100]}', 0, 1)
        self.set_text_color(0, 0, 0)

    def columna(self, nombre: str, tipo: str, muestra: str):
        self.set_font('Courier', '', 7)
        self.cell(0, 4,
                  f'    - {_clean(nombre)[:30]} [{tipo}]: {_clean(muestra)[:55]}',
                  0, 1)

    def tabla_muestra(self, titulo: str, filas: List[Dict], columnas: List[str]):
        """Escribe una tabla de Head o Tail en el PDF."""
        if not filas:
            return
        self.set_font('Arial', 'B', 8)
        self.cell(0, 5, f'  {titulo}:', 0, 1)
        self.set_font('Courier', '', 6)
        # Encabezado de tabla
        header_line = " | ".join(str(c)[:15] for c in columnas[:8])
        self.cell(0, 4, f'    {_clean(header_line)[:130]}', 0, 1)
        self.cell(0, 3, '    ' + '-' * 100, 0, 1)
        for fila in filas:
            vals = " | ".join(str(fila.get(c, ''))[:15] for c in columnas[:8])
            self.cell(0, 4, f'    {_clean(vals)[:130]}', 0, 1)


def _clean(txt: str) -> str:
    return str(txt).encode('latin-1', 'replace').decode('latin-1')


# ─────────────────────────────────────────────
# Detección inteligente de encabezados
# ─────────────────────────────────────────────

_PALABRAS_HEADER = {
    'id', 'nombre', 'name', 'fecha', 'date', 'codigo', 'code', 'tipo', 'type',
    'descripcion', 'description', 'cantidad', 'amount', 'total', 'precio', 'price',
    'estado', 'status', 'usuario', 'user', 'email', 'telefono', 'phone',
    'direccion', 'address', 'ciudad', 'city', 'pais', 'country', 'numero',
    'clave', 'key', 'valor', 'value', 'categoria', 'category', 'producto',
    'cliente', 'customer', 'orden', 'order', 'factura', 'invoice', 'cuenta',
    'año', 'year', 'mes', 'month', 'dia', 'day', 'hora', 'time',
    'created', 'updated', 'grupo', 'actividad', 'area', 'region',
}


def _puntuar_fila_header(fila: List) -> Tuple[bool, float]:
    if not fila:
        return False, 0.0
    celdas_no_vacias = [c for c in fila if c and str(c).strip()]
    if len(celdas_no_vacias) < 2:
        return False, 0.0

    score = 0.0
    for celda in celdas_no_vacias:
        s = str(celda).strip().lower()
        try:
            float(s.replace(',', '.'))
            score -= 0.3
            continue
        except ValueError:
            pass
        for palabra in _PALABRAS_HEADER:
            if palabra in s:
                score += 0.5
                break
        if re.match(r'^[a-zA-Z][a-zA-Z0-9_ ]*$', s):
            score += 0.3
        if 2 <= len(s) <= 50:
            score += 0.2
        if len(s) > 100:
            score -= 0.5

    score /= len(celdas_no_vacias)
    if len(celdas_no_vacias) / len(fila) < 0.5:
        score -= 0.5

    return score > 0.1, score


def _buscar_encabezado(filas: List[List], callback=None) -> Tuple[int, List[str]]:
    mejor_score, mejor_idx, mejor_fila = -1, 0, []
    for idx, fila in enumerate(filas[:MAX_FILAS_BUSQUEDA_HEADER]):
        es_ok, sc = _puntuar_fila_header(fila)
        if es_ok and sc > mejor_score:
            mejor_score, mejor_idx, mejor_fila = sc, idx, fila

    if callback and mejor_idx > 0:
        callback(f"   Encabezados detectados en fila {mejor_idx + 1}")

    if not mejor_fila and filas:
        mejor_fila, mejor_idx = filas[0], 0

    headers = [
        str(h).strip() if (h and str(h).strip()) else f"Col_{i+1}"
        for i, h in enumerate(mejor_fila)
    ]
    return mejor_idx, headers


# ─────────────────────────────────────────────
# Analizadores por formato
# ─────────────────────────────────────────────

def _analizar_csv(ruta: str, callback=None) -> Dict:
    import csv

    res = {
        'tipo': 'CSV', 'ruta': ruta,
        'tamaño_bytes': os.path.getsize(ruta),
        'columnas': [], 'tipos_detectados': {},
        'muestra_valores': {}, 'fila_encabezado': 1,
        'head': [], 'tail': [], 'total_filas': 0,
    }

    try:
        with open(ruta, 'r', encoding='utf-8', errors='replace', newline='') as f:
            muestra_raw = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(muestra_raw)
                reader = csv.reader(f, dialect)
            except Exception:
                reader = csv.reader(f)

            primeras: List[List] = []
            for i, row in enumerate(reader):
                primeras.append(row)
                if i >= MAX_FILAS_BUSQUEDA_HEADER:
                    break

            if not primeras:
                return res

            idx_h, headers = _buscar_encabezado(primeras, callback)
            res['fila_encabezado'] = idx_h + 1
            res['columnas'] = headers

            valores_unicos = {c: set() for c in headers}
            res['tipos_detectados'] = {c: 'texto' for c in headers}

            todas_filas: List[List] = list(primeras[idx_h + 1:])

            for row in reader:
                todas_filas.append(row)
                if callback and len(todas_filas) % CHUNK_SIZE == 0:
                    callback(f"CSV: {len(todas_filas):,} filas leídas...")

            res['total_filas'] = len(todas_filas)

            # Head & Tail
            res['head'] = _filas_a_dicts(todas_filas[:HEAD_TAIL_ROWS], headers)
            res['tail'] = _filas_a_dicts(todas_filas[-HEAD_TAIL_ROWS:], headers)

            # Muestra y tipos
            for fnum, row in enumerate(todas_filas):
                for i, val in enumerate(row):
                    if i >= len(headers):
                        break
                    col = headers[i]
                    if val and str(val).strip():
                        if len(valores_unicos[col]) < MUESTRA_VALORES:
                            valores_unicos[col].add(str(val)[:50])
                        if fnum < 1000:
                            t = detectar_tipo_str(str(val))
                            if (res['tipos_detectados'][col] == 'texto'
                                    and t != 'texto'):
                                res['tipos_detectados'][col] = t

        res['muestra_valores'] = {k: list(v) for k, v in valores_unicos.items()}

    except Exception as exc:
        res['error'] = str(exc)

    return res


def _analizar_excel(ruta: str, callback=None) -> Dict:
    try:
        import openpyxl
    except ImportError:
        return {'error': 'Instala openpyxl: pip install openpyxl', 'ruta': ruta}

    res = {
        'tipo': 'Excel', 'ruta': ruta,
        'tamaño_bytes': os.path.getsize(ruta),
        'hojas': [],
    }

    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        for nombre_hoja in wb.sheetnames:
            if callback:
                callback(f"Excel: procesando hoja '{nombre_hoja}'...")

            ws = wb[nombre_hoja]
            hoja: Dict = {
                'nombre': nombre_hoja,
                'columnas': [], 'tipos_detectados': {},
                'muestra_valores': {}, 'fila_encabezado': 1,
                'head': [], 'tail': [], 'total_filas': 0,
            }

            todas_filas: List[List] = [list(row) for row in ws.iter_rows(values_only=True)]
            primeras = todas_filas[:MAX_FILAS_BUSQUEDA_HEADER]

            if not primeras:
                res['hojas'].append(hoja)
                continue

            idx_h, headers = _buscar_encabezado(primeras, callback)
            hoja['fila_encabezado'] = idx_h + 1
            hoja['columnas'] = headers

            datos = todas_filas[idx_h + 1:]
            hoja['total_filas'] = len(datos)
            hoja['head'] = _filas_a_dicts(datos[:HEAD_TAIL_ROWS], headers)
            hoja['tail'] = _filas_a_dicts(datos[-HEAD_TAIL_ROWS:], headers)

            valores_unicos = {c: set() for c in headers}
            hoja['tipos_detectados'] = {c: 'texto' for c in headers}

            for fnum, row in enumerate(datos):
                for i, val in enumerate(row):
                    if i >= len(headers):
                        break
                    col = headers[i]
                    if val is not None:
                        if len(valores_unicos[col]) < MUESTRA_VALORES:
                            valores_unicos[col].add(str(val)[:50])
                        if fnum < 1000:
                            t = detectar_tipo_valor(val)
                            if (hoja['tipos_detectados'][col] == 'texto'
                                    and t != 'texto'):
                                hoja['tipos_detectados'][col] = t

            hoja['muestra_valores'] = {k: list(v) for k, v in valores_unicos.items()}
            res['hojas'].append(hoja)

        wb.close()

    except Exception as exc:
        res['error'] = str(exc)

    return res


def _analizar_parquet(ruta: str, callback=None) -> Dict:
    try:
        import pyarrow.parquet as pq
    except ImportError:
        return {'error': 'Instala pyarrow: pip install pyarrow', 'ruta': ruta}

    res = {
        'tipo': 'Parquet', 'ruta': ruta,
        'tamaño_bytes': os.path.getsize(ruta),
        'columnas': [], 'tipos_detectados': {},
        'muestra_valores': {}, 'total_filas': 0,
        'head': [], 'tail': [],
    }

    try:
        if callback:
            callback("Parquet: leyendo metadata...")

        pf = pq.ParquetFile(ruta)
        schema = pf.schema_arrow
        res['total_filas'] = pf.metadata.num_rows
        res['columnas'] = [f.name for f in schema]
        res['tipos_detectados'] = {
            f.name: _tipo_arrow(str(f.type)) for f in schema
        }

        valores_unicos = {c: set() for c in res['columnas']}

        try:
            for batch in pf.iter_batches(batch_size=min(500, CHUNK_SIZE)):
                for col_name in res['columnas']:
                    if len(valores_unicos[col_name]) >= MUESTRA_VALORES:
                        continue
                    try:
                        col = batch.column(batch.schema.get_field_index(col_name))
                        for i in range(min(len(col), MUESTRA_VALORES * 2)):
                            v = col[i].as_py()
                            if v is not None:
                                valores_unicos[col_name].add(str(v)[:50])
                    except Exception:
                        pass
                break  # Solo primer batch para muestra
        except Exception as exc2:
            if callback:
                callback(f"Parquet: no se pudo extraer muestra ({exc2})")

        # Head & Tail usando pandas (más sencillo para parquet)
        try:
            import pandas as pd
            df_head = pd.read_parquet(ruta).head(HEAD_TAIL_ROWS)
            df_tail = pd.read_parquet(ruta).tail(HEAD_TAIL_ROWS)
            res['head'] = df_head.to_dict(orient='records')
            res['tail'] = df_tail.to_dict(orient='records')
        except Exception:
            pass  # No crítico

        res['muestra_valores'] = {k: list(v) for k, v in valores_unicos.items()}

    except Exception as exc:
        res['error'] = str(exc)

    return res


def _tipo_arrow(t: str) -> str:
    t = t.lower()
    if 'int' in t:
        return 'entero'
    if 'float' in t or 'double' in t:
        return 'decimal'
    if 'timestamp' in t or 'date' in t:
        return 'fecha'
    if 'bool' in t:
        return 'booleano'
    if 'string' in t or 'utf8' in t:
        return 'texto'
    return t


def _filas_a_dicts(filas: List[List], headers: List[str]) -> List[Dict]:
    result = []
    for fila in filas:
        d = {}
        for i, h in enumerate(headers):
            d[h] = fila[i] if i < len(fila) else ''
        result.append(d)
    return result


# ─────────────────────────────────────────────
# Función principal: generar informe
# ─────────────────────────────────────────────

def generar_informe_datos(
    ruta_directorio: str,
    carpeta_salida: Optional[str] = None,
    nombre_informe: str = "mapa_datos.txt",
    formato: str = 'txt',
    callback: Optional[Callable[[str], None]] = None,
) -> Dict:
    """
    Escanea un directorio buscando archivos de datos y genera un informe
    con estructura, tipos y muestras Head/Tail.

    Args:
        ruta_directorio: Carpeta a escanear.
        carpeta_salida:  Carpeta de salida (None = misma que origen).
        nombre_informe:  Nombre base del informe (extensión se ajusta al formato).
        formato:         'txt' (defecto) | 'pdf' | 'csv'.
        callback:        Función de progreso.

    Returns:
        Dict con archivos_analizados, ruta_reporte, resultados.
    """
    def log(msg):
        if callback:
            callback(msg)
        print(msg)

    ruta_directorio = os.path.normpath(os.path.abspath(ruta_directorio))
    carpeta_salida = (
        os.path.normpath(os.path.abspath(carpeta_salida))
        if carpeta_salida else ruta_directorio
    )

    log(f"Buscando archivos de datos en: {ruta_directorio}")

    archivos = []
    for root, dirs, files in os.walk(ruta_directorio):
        dirs[:] = [d for d in dirs
                   if d not in {'.git', '__pycache__', 'node_modules',
                                'venv', '.venv', 'REFACTORIZACION'}]
        for f in files:
            if os.path.splitext(f)[1].lower() in EXTENSIONES_DATOS:
                archivos.append(os.path.join(root, f))

    log(f"Encontrados: {len(archivos)} archivos de datos")

    resultados = []
    for i, ruta in enumerate(archivos):
        ext = os.path.splitext(ruta)[1].lower()
        log(f"[{i+1}/{len(archivos)}] Analizando: {os.path.basename(ruta)}")
        try:
            if ext == '.csv':
                r = _analizar_csv(ruta, callback)
            elif ext in ('.xlsx', '.xls', '.xlsm'):
                r = _analizar_excel(ruta, callback)
            elif ext in ('.parquet', '.pq'):
                r = _analizar_parquet(ruta, callback)
            else:
                continue
            resultados.append(r)
        except Exception as exc:
            log(f"  Error: {exc}")
            resultados.append({'ruta': ruta, 'error': str(exc)})

    os.makedirs(carpeta_salida, exist_ok=True)

    # Ajustar extensión
    base = os.path.splitext(nombre_informe)[0]
    fmt = formato.lower()
    nombre_final = f"{base}.{'txt' if fmt == 'txt' else fmt}"
    ruta_salida = os.path.join(carpeta_salida, nombre_final)

    try:
        if fmt == 'txt':
            _escribir_txt(resultados, ruta_salida, log)
        elif fmt == 'pdf':
            _escribir_pdf(resultados, ruta_salida, log)
        elif fmt == 'csv':
            _escribir_csv_informe(resultados, ruta_salida, log)
    except Exception as exc:
        log(f"❌ Error generando informe {fmt.upper()}: {exc}")
        ruta_salida = None

    return {
        'archivos_analizados': len(resultados),
        'ruta_reporte': ruta_salida,
        'resultados': resultados,
    }


# ─────────────────────────────────────────────
# Writers de informe
# ─────────────────────────────────────────────

def _escribir_txt(resultados: List[Dict], ruta: str, log):
    """Genera el informe en formato TXT estructurado con Head & Tail."""
    sep_mayor = "=" * 70
    sep_menor = "-" * 50

    with open(ruta, 'w', encoding='utf-8') as f:
        f.write("MAPA DE ARCHIVOS DE DATOS\n")
        f.write(f"{sep_mayor}\n\n")

        for res in resultados:
            nombre = os.path.basename(res.get('ruta', 'N/A'))
            f.write(f"ARCHIVO : {nombre}\n")
            f.write(f"Tipo    : {res.get('tipo', 'Desconocido')}\n")
            f.write(f"Ruta    : {res.get('ruta', 'N/A')}\n")
            f.write(f"Tamaño  : {formato_bytes(res.get('tamaño_bytes', 0))}\n")

            if 'error' in res:
                f.write(f"ERROR   : {res['error']}\n\n")
                f.write(sep_menor + "\n\n")
                continue

            def escribir_bloque(datos: Dict, indent: str = ""):
                cols = datos.get('columnas', [])
                tipos = datos.get('tipos_detectados', {})
                muestras = datos.get('muestra_valores', {})
                f.write(f"{indent}Encabezado en fila : {datos.get('fila_encabezado', 1)}\n")
                f.write(f"{indent}Total filas        : {datos.get('total_filas', 0):,}\n")
                f.write(f"{indent}Columnas ({len(cols)}):\n")
                for col in cols:
                    t = tipos.get(col, 'texto')
                    m = ", ".join(list(muestras.get(col, []))[:3]) or "(vacío)"
                    if len(m) > 60:
                        m = m[:57] + "..."
                    f.write(f"{indent}  - {col} [{t}] : {m}\n")

                # Head
                head = datos.get('head', [])
                if head:
                    f.write(f"\n{indent}HEAD (primeras {len(head)} filas):\n")
                    f.write(f"{indent}  " + " | ".join(str(c)[:14] for c in cols[:8]) + "\n")
                    f.write(f"{indent}  " + "-" * 60 + "\n")
                    for fila in head:
                        vals = " | ".join(str(fila.get(c, ''))[:14] for c in cols[:8])
                        f.write(f"{indent}  {vals}\n")

                # Tail
                tail = datos.get('tail', [])
                if tail:
                    f.write(f"\n{indent}TAIL (últimas {len(tail)} filas):\n")
                    f.write(f"{indent}  " + " | ".join(str(c)[:14] for c in cols[:8]) + "\n")
                    f.write(f"{indent}  " + "-" * 60 + "\n")
                    for fila in tail:
                        vals = " | ".join(str(fila.get(c, ''))[:14] for c in cols[:8])
                        f.write(f"{indent}  {vals}\n")

            if 'hojas' in res:
                for hoja in res['hojas']:
                    f.write(f"\n  HOJA: {hoja['nombre']}\n")
                    escribir_bloque(hoja, indent="  ")
            else:
                escribir_bloque(res)

            f.write(f"\n{sep_menor}\n\n")

        f.write(f"Total archivos analizados: {len(resultados)}\n")

    if os.path.exists(ruta):
        log(f"✅ Informe TXT guardado: {ruta}")
    else:
        log(f"❌ Error: Informe no generado en {ruta}")


def _escribir_pdf(resultados: List[Dict], ruta: str, log):
    """Genera el informe en formato PDF profesional con Head & Tail."""
    pdf = _PDFDatos()
    pdf.add_page()

    for res in resultados:
        nombre = os.path.basename(res.get('ruta', 'N/A'))
        pdf.archivo_header(nombre, res.get('tipo', '?'), res.get('ruta', 'N/A'))

        if 'error' in res:
            pdf.error_msg(res['error'])
            pdf.ln(3)
            continue

        pdf.info(f"Tamaño: {formato_bytes(res.get('tamaño_bytes', 0))}")

        def escribir_bloque_pdf(datos: Dict):
            cols = datos.get('columnas', [])
            tipos = datos.get('tipos_detectados', {})
            muestras = datos.get('muestra_valores', {})
            pdf.info(f"Encabezado fila: {datos.get('fila_encabezado', 1)} | "
                     f"Filas: {datos.get('total_filas', 0):,} | "
                     f"Columnas: {len(cols)}")
            for col in cols:
                t = tipos.get(col, 'texto')
                m_list = list(muestras.get(col, []))[:3]
                m = ", ".join(m_list) if m_list else "(vacío)"
                if len(m) > 55:
                    m = m[:52] + "..."
                pdf.columna(col, t, m)
            if datos.get('head'):
                pdf.tabla_muestra(f"HEAD ({HEAD_TAIL_ROWS} filas)", datos['head'], cols)
            if datos.get('tail'):
                pdf.tabla_muestra(f"TAIL ({HEAD_TAIL_ROWS} filas)", datos['tail'], cols)

        if 'hojas' in res:
            for hoja in res['hojas']:
                pdf.hoja_header(hoja['nombre'])
                escribir_bloque_pdf(hoja)
        else:
            escribir_bloque_pdf(res)

        pdf.ln(3)

    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 8, f"Total archivos analizados: {len(resultados)}", 0, 1)
    pdf.output(ruta)

    if os.path.exists(ruta):
        compress_pdf(ruta, callback=log)
        log(f"✅ Informe PDF guardado: {ruta}")
    else:
        log(f"❌ Error: Informe PDF no generado en {ruta}")


def _escribir_csv_informe(resultados: List[Dict], ruta: str, log):
    """Genera el informe en formato CSV (flat)."""
    import csv

    campos = [
        'archivo_nombre', 'archivo_ruta', 'archivo_tipo', 'archivo_tamaño',
        'hoja_nombre', 'fila_encabezado', 'total_filas', 'total_columnas',
        'columna_nombre', 'columna_tipo', 'ejemplo_valores', 'head_muestra',
        'error',
    ]

    with open(ruta, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=campos)
        writer.writeheader()

        for res in resultados:
            base = {
                'archivo_nombre': os.path.basename(res.get('ruta', '')),
                'archivo_ruta': res.get('ruta', ''),
                'archivo_tipo': res.get('tipo', ''),
                'archivo_tamaño': formato_bytes(res.get('tamaño_bytes', 0)),
                'error': res.get('error', ''),
            }
            if 'error' in res:
                writer.writerow(base)
                continue

            def escribir_cols_csv(datos: Dict, hoja_nom: str = ''):
                cols = datos.get('columnas', [])
                tipos = datos.get('tipos_detectados', {})
                muestras = datos.get('muestra_valores', {})
                head = datos.get('head', [])
                head_str = str(head[:2]) if head else ''

                for col in cols:
                    row = base.copy()
                    row.update({
                        'hoja_nombre': hoja_nom,
                        'fila_encabezado': datos.get('fila_encabezado', 1),
                        'total_filas': datos.get('total_filas', 0),
                        'total_columnas': len(cols),
                        'columna_nombre': col,
                        'columna_tipo': tipos.get(col, 'texto'),
                        'ejemplo_valores': "; ".join(list(muestras.get(col, []))[:5]),
                        'head_muestra': head_str[:200],
                    })
                    writer.writerow(row)

            if 'hojas' in res:
                for hoja in res['hojas']:
                    escribir_cols_csv(hoja, hoja['nombre'])
            else:
                escribir_cols_csv(res)

    if os.path.exists(ruta):
        log(f"✅ Informe CSV guardado: {ruta}")
    else:
        log(f"❌ Error: Informe CSV no generado en {ruta}")
