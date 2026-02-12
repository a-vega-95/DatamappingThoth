"""
Analizador de Archivos de Datos
Mapea estructura y contenido de CSV, Excel, Parquet
Lectura eficiente por chunks para no saturar RAM
Detección inteligente de encabezados
"""

import os
import re
from typing import Dict, List, Optional, Callable, Tuple
from fpdf import FPDF

# Tamaño de chunk para lectura eficiente
CHUNK_SIZE = 10000  # filas por chunk
MUESTRA_VALORES = 5  # valores únicos de muestra por columna
MAX_FILAS_BUSQUEDA_HEADER = 20  # Máximo de filas a analizar para encontrar encabezados


class PDFDatos(FPDF):
    """Clase para generar reportes de datos en PDF"""
    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'Mapa de Archivos de Datos', 0, 1, 'C')
        self.set_font('Arial', 'I', 8)
        self.cell(0, 5, 'Deteccion inteligente de encabezados habilitada', 0, 1, 'C')
        self.ln(2)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Pagina {self.page_no()}', 0, 0, 'C')
    
    def archivo_header(self, nombre, tipo, ruta):
        self.set_fill_color(200, 220, 255)
        self.set_font('Arial', 'B', 10)
        nombre_clean = nombre.encode('latin-1', 'replace').decode('latin-1')
        self.cell(0, 7, f'ARCHIVO: {nombre_clean}', 0, 1, 'L', 1)
        self.set_font('Arial', '', 8)
        self.cell(0, 5, f'Tipo: {tipo}', 0, 1)
        ruta_clean = ruta.encode('latin-1', 'replace').decode('latin-1')[:100]
        self.cell(0, 5, f'Ruta: {ruta_clean}', 0, 1)
    
    def error(self, mensaje):
        self.set_font('Arial', 'I', 8)
        self.set_text_color(255, 0, 0)
        self.cell(0, 5, f'ERROR: {mensaje[:80]}', 0, 1)
        self.set_text_color(0, 0, 0)
    
    def info(self, texto):
        self.set_font('Arial', '', 8)
        texto_clean = texto.encode('latin-1', 'replace').decode('latin-1')
        self.cell(0, 5, texto_clean[:120], 0, 1)
    
    def hoja_header(self, nombre):
        self.set_font('Arial', 'B', 9)
        self.set_fill_color(230, 230, 230)
        nombre_clean = nombre.encode('latin-1', 'replace').decode('latin-1')
        self.cell(0, 6, f'  HOJA: {nombre_clean}', 0, 1, 'L', 1)
    
    def columna(self, nombre, tipo, muestra):
        self.set_font('Courier', '', 7)
        nombre_clean = nombre.encode('latin-1', 'replace').decode('latin-1')[:30]
        muestra_clean = muestra.encode('latin-1', 'replace').decode('latin-1')[:50]
        self.cell(0, 4, f'    - {nombre_clean} [{tipo}]: {muestra_clean}', 0, 1)


def _es_encabezado_valido(fila: List, min_columnas: int = 2) -> Tuple[bool, float]:
    """
    Evalúa si una fila parece ser un encabezado.
    Retorna (es_encabezado, score)
    
    Criterios:
    - Textos descriptivos (no números puros)
    - Sin valores vacíos excesivos
    - Nombres de columna típicos
    - Longitud razonable de texto
    """
    if not fila or len([c for c in fila if c]) < min_columnas:
        return False, 0.0
    
    score = 0.0
    num_celdas = len(fila)
    celdas_validas = 0
    
    # Palabras clave comunes en encabezados
    palabras_header = {
        'id', 'nombre', 'name', 'fecha', 'date', 'codigo', 'code', 'tipo', 'type',
        'descripcion', 'description', 'cantidad', 'amount', 'total', 'precio', 'price',
        'estado', 'status', 'usuario', 'user', 'email', 'telefono', 'phone',
        'direccion', 'address', 'ciudad', 'city', 'pais', 'country', 'numero', 'number',
        'clave', 'key', 'valor', 'value', 'categoria', 'category', 'producto', 'product',
        'cliente', 'customer', 'orden', 'order', 'factura', 'invoice', 'cuenta', 'account',
        'año', 'year', 'mes', 'month', 'dia', 'day', 'hora', 'time', 'created', 'updated'
    }
    
    for celda in fila:
        if celda is None or str(celda).strip() == '':
            continue
        
        celda_str = str(celda).strip().lower()
        celdas_validas += 1
        
        # Penalizar si es un número puro
        try:
            float(celda_str.replace(',', '.'))
            score -= 0.3
            continue
        except:
            pass
        
        # Bonus si contiene palabras clave de encabezado
        for palabra in palabras_header:
            if palabra in celda_str:
                score += 0.5
                break
        
        # Bonus si tiene formato de identificador (snake_case, camelCase, etc.)
        if re.match(r'^[a-zA-Z][a-zA-Z0-9_]*$', celda_str):
            score += 0.3
        
        # Bonus si tiene longitud razonable para un título
        if 2 <= len(celda_str) <= 50:
            score += 0.2
        
        # Penalizar textos muy largos (probablemente datos)
        if len(celda_str) > 100:
            score -= 0.5
    
    # Normalizar score
    if celdas_validas > 0:
        score = score / celdas_validas
    
    # Verificar proporción de celdas válidas
    ratio_validas = celdas_validas / num_celdas if num_celdas > 0 else 0
    if ratio_validas < 0.5:
        score -= 0.5
    
    return score > 0.1, score


def _buscar_fila_encabezado(filas: List[List], callback: Optional[Callable] = None) -> Tuple[int, List]:
    """
    Busca inteligentemente la fila que contiene los encabezados.
    Retorna (índice_fila, encabezados)
    """
    mejor_score = -1
    mejor_idx = 0
    mejor_fila = []
    
    for idx, fila in enumerate(filas[:MAX_FILAS_BUSQUEDA_HEADER]):
        es_header, score = _es_encabezado_valido(fila)
        
        if es_header and score > mejor_score:
            mejor_score = score
            mejor_idx = idx
            mejor_fila = fila
    
    if callback and mejor_idx > 0:
        callback(f"   Encabezados detectados en fila {mejor_idx + 1}")
    
    # Si no encontramos un buen header, usar la primera fila
    if not mejor_fila and filas:
        mejor_fila = filas[0]
        mejor_idx = 0
    
    # Limpiar encabezados
    headers = []
    for i, h in enumerate(mejor_fila):
        if h is None or str(h).strip() == '':
            headers.append(f"Col_{i+1}")
        else:
            headers.append(str(h).strip())
    
    return mejor_idx, headers


def analizar_csv(ruta: str, callback: Optional[Callable] = None) -> Dict:
    """Analiza archivo CSV por chunks con detección inteligente de encabezados"""
    import csv
    
    resultado = {
        'tipo': 'CSV',
        'ruta': ruta,
        'columnas': [],
        'total_filas': 0,
        'tamaño_bytes': os.path.getsize(ruta),
        'muestra_valores': {},
        'tipos_detectados': {},
        'fila_encabezado': 1
    }
    
    valores_unicos = {}  # Para trackear valores únicos por columna
    
    try:
        with open(ruta, 'r', encoding='utf-8', errors='replace', newline='') as f:
            # Detectar delimitador
            sample = f.read(8192)
            f.seek(0)
            
            try:
                dialect = csv.Sniffer().sniff(sample)
                reader = csv.reader(f, dialect)
            except:
                reader = csv.reader(f)
            
            # Leer primeras filas para detectar encabezados
            primeras_filas = []
            for i, row in enumerate(reader):
                primeras_filas.append(row)
                if i >= MAX_FILAS_BUSQUEDA_HEADER:
                    break
            
            if not primeras_filas:
                return resultado
            
            # Buscar encabezados inteligentemente
            idx_header, headers = _buscar_fila_encabezado(primeras_filas, callback)
            resultado['fila_encabezado'] = idx_header + 1
            resultado['columnas'] = headers
            
            for col in headers:
                valores_unicos[col] = set()
                resultado['tipos_detectados'][col] = 'texto'
            
            # Procesar filas después del encabezado (de las primeras leídas)
            filas_leidas = 0
            for row in primeras_filas[idx_header + 1:]:
                filas_leidas += 1
                _procesar_fila_datos(row, headers, valores_unicos, resultado, filas_leidas)
            
            # Continuar leyendo el resto del archivo
            for row in reader:
                filas_leidas += 1
                _procesar_fila_datos(row, headers, valores_unicos, resultado, filas_leidas)
                
                if callback and filas_leidas % CHUNK_SIZE == 0:
                    callback(f"CSV: {filas_leidas:,} filas procesadas...")
            
            resultado['total_filas'] = filas_leidas
            resultado['muestra_valores'] = {k: list(v) for k, v in valores_unicos.items()}
            
    except Exception as e:
        resultado['error'] = str(e)
    
    return resultado


def _procesar_fila_datos(row: List, headers: List, valores_unicos: Dict, 
                         resultado: Dict, fila_num: int):
    """Procesa una fila de datos actualizando estadísticas"""
    for i, val in enumerate(row):
        if i < len(headers):
            col = headers[i]
            if val and str(val).strip():
                # Muestrear valores
                if len(valores_unicos[col]) < MUESTRA_VALORES:
                    valores_unicos[col].add(str(val)[:50])
                
                # Detectar tipos en primeras filas
                if fila_num <= 1000:
                    tipo = _detectar_tipo(str(val))
                    if resultado['tipos_detectados'][col] == 'texto' and tipo != 'texto':
                        resultado['tipos_detectados'][col] = tipo


def analizar_excel(ruta: str, callback: Optional[Callable] = None) -> Dict:
    """Analiza archivo Excel con detección inteligente de encabezados"""
    try:
        import openpyxl
    except ImportError:
        return {'error': 'Instala openpyxl: pip install openpyxl', 'ruta': ruta}
    
    resultado = {
        'tipo': 'Excel',
        'ruta': ruta,
        'hojas': [],
        'tamaño_bytes': os.path.getsize(ruta)
    }
    
    try:
        # read_only=True para eficiencia de memoria
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        
        for nombre_hoja in wb.sheetnames:
            if callback:
                callback(f"Excel: Analizando hoja '{nombre_hoja}'...")
            
            ws = wb[nombre_hoja]
            hoja_info = {
                'nombre': nombre_hoja,
                'columnas': [],
                'total_filas': 0,
                'muestra_valores': {},
                'tipos_detectados': {},
                'fila_encabezado': 1
            }
            
            # Leer primeras filas para detectar encabezados
            primeras_filas = []
            todas_filas = []
            
            for row in ws.iter_rows(values_only=True):
                todas_filas.append(list(row))
                if len(primeras_filas) < MAX_FILAS_BUSQUEDA_HEADER:
                    primeras_filas.append(list(row))
            
            if not primeras_filas:
                resultado['hojas'].append(hoja_info)
                continue
            
            # Buscar encabezados inteligentemente
            idx_header, headers = _buscar_fila_encabezado(primeras_filas, callback)
            hoja_info['fila_encabezado'] = idx_header + 1
            hoja_info['columnas'] = headers
            
            valores_unicos = {col: set() for col in headers}
            hoja_info['tipos_detectados'] = {col: 'texto' for col in headers}
            
            # Procesar datos después del encabezado
            filas_datos = 0
            for row in todas_filas[idx_header + 1:]:
                filas_datos += 1
                for i, val in enumerate(row):
                    if i < len(headers):
                        col = headers[i]
                        if val is not None:
                            if len(valores_unicos[col]) < MUESTRA_VALORES:
                                valores_unicos[col].add(str(val)[:50])
                            
                            if filas_datos <= 1000:
                                tipo = _detectar_tipo_valor(val)
                                if hoja_info['tipos_detectados'][col] == 'texto' and tipo != 'texto':
                                    hoja_info['tipos_detectados'][col] = tipo
                
                if callback and filas_datos % CHUNK_SIZE == 0:
                    callback(f"Excel [{nombre_hoja}]: {filas_datos:,} filas...")
            
            hoja_info['total_filas'] = filas_datos
            hoja_info['muestra_valores'] = {k: list(v) for k, v in valores_unicos.items()}
            resultado['hojas'].append(hoja_info)
        
        wb.close()
        
    except Exception as e:
        resultado['error'] = str(e)
    
    return resultado


def analizar_parquet(ruta: str, callback: Optional[Callable] = None) -> Dict:
    """Analiza archivo Parquet usando solo PyArrow (sin pandas)"""
    try:
        import pyarrow.parquet as pq
        import pyarrow as pa
    except ImportError:
        return {'error': 'Instala pyarrow: pip install pyarrow', 'ruta': ruta}
    
    resultado = {
        'tipo': 'Parquet',
        'ruta': ruta,
        'columnas': [],
        'total_filas': 0,
        'tamaño_bytes': os.path.getsize(ruta),
        'muestra_valores': {},
        'tipos_detectados': {}
    }
    
    try:
        if callback:
            callback("Parquet: Leyendo metadata...")
        
        # Leer solo metadata primero (muy eficiente)
        parquet_file = pq.ParquetFile(ruta)
        metadata = parquet_file.metadata
        schema = parquet_file.schema_arrow
        
        resultado['total_filas'] = metadata.num_rows
        resultado['columnas'] = [field.name for field in schema]
        resultado['tipos_detectados'] = {field.name: _traducir_tipo_arrow(str(field.type)) for field in schema}
        
        # Leer muestra de valores usando PyArrow puro (sin pandas)
        if callback:
            callback("Parquet: Extrayendo muestra de valores...")
        
        valores_unicos = {col: set() for col in resultado['columnas']}
        
        # Leer solo las primeras filas para muestra (más eficiente)
        try:
            # Leer un batch pequeño
            for batch in parquet_file.iter_batches(batch_size=min(1000, CHUNK_SIZE)):
                for col_name in resultado['columnas']:
                    if len(valores_unicos[col_name]) >= MUESTRA_VALORES:
                        continue
                    
                    try:
                        col_idx = batch.schema.get_field_index(col_name)
                        column = batch.column(col_idx)
                        
                        # Extraer valores como Python nativos
                        for i in range(min(len(column), MUESTRA_VALORES * 2)):
                            if len(valores_unicos[col_name]) >= MUESTRA_VALORES:
                                break
                            val = column[i].as_py()
                            if val is not None:
                                valores_unicos[col_name].add(str(val)[:50])
                    except Exception:
                        continue
                
                # Solo necesitamos un batch para muestra
                break
                
        except Exception as e:
            if callback:
                callback(f"Parquet: Nota - no se pudo extraer muestra: {str(e)[:50]}")
        
        resultado['muestra_valores'] = {k: list(v) for k, v in valores_unicos.items()}
        
    except Exception as e:
        resultado['error'] = str(e)
    
    return resultado


def _traducir_tipo_arrow(tipo_arrow: str) -> str:
    """Traduce tipos de PyArrow a nombres más legibles"""
    tipo_lower = tipo_arrow.lower()
    
    if 'int' in tipo_lower:
        return 'entero'
    elif 'float' in tipo_lower or 'double' in tipo_lower:
        return 'decimal'
    elif 'timestamp' in tipo_lower or 'date' in tipo_lower:
        return 'fecha'
    elif 'bool' in tipo_lower:
        return 'booleano'
    elif 'string' in tipo_lower or 'utf8' in tipo_lower:
        return 'texto'
    else:
        return tipo_arrow  # Mantener el original si no se reconoce


def _detectar_tipo(valor_str: str) -> str:
    """Detecta el tipo de dato de un string"""
    if not valor_str or valor_str.strip() == '':
        return 'texto'
    
    # Intentar número entero
    try:
        int(valor_str.replace(',', '').replace('.', '').strip())
        if '.' in valor_str or ',' in valor_str:
            return 'decimal'
        return 'entero'
    except:
        pass
    
    # Intentar decimal
    try:
        float(valor_str.replace(',', '.').strip())
        return 'decimal'
    except:
        pass
    
    # Intentar fecha
    if any(sep in valor_str for sep in ['/', '-']) and len(valor_str) <= 20:
        if any(c.isdigit() for c in valor_str):
            return 'fecha'
    
    return 'texto'


def _detectar_tipo_valor(valor) -> str:
    """Detecta el tipo de un valor Python"""
    if valor is None:
        return 'texto'
    
    tipo = type(valor).__name__
    
    if tipo in ('int', 'int64', 'int32'):
        return 'entero'
    elif tipo in ('float', 'float64', 'float32'):
        return 'decimal'
    elif tipo == 'datetime':
        return 'fecha'
    elif tipo == 'bool':
        return 'booleano'
    
    return 'texto'


def generar_reporte_datos(ruta_directorio: str, archivo_salida: str = "mapa_datos",
                          carpeta_salida: Optional[str] = None,
                          callback: Optional[Callable] = None,
                          formato: str = 'pdf') -> Dict:
    """
    Escanea directorio buscando archivos de datos y genera reporte.
    
    Args:
        ruta_directorio: Carpeta a escanear
        archivo_salida: Nombre base del archivo de reporte
        carpeta_salida: Carpeta donde guardar el reporte (None = misma que ruta_directorio)
        callback: Función para reportar progreso
        formato: Formato de salida ('pdf', 'txt', 'csv')
    
    Returns:
        Dict con resumen del análisis
    """
    extensiones_datos = {'.csv', '.xlsx', '.xls', '.parquet', '.pq'}
    archivos_encontrados = []
    resultados = []
    
    def log(msg):
        if callback:
            callback(msg)
        print(msg)
    
    # 0. Normalizar rutas
    ruta_directorio = os.path.normpath(os.path.abspath(ruta_directorio))
    if carpeta_salida:
        carpeta_salida = os.path.normpath(os.path.abspath(carpeta_salida))
    else:
        carpeta_salida = ruta_directorio

    log(f"Buscando archivos de datos en: {ruta_directorio}")
    
    # Buscar archivos
    for root, dirs, files in os.walk(ruta_directorio):
        # Ignorar carpetas comunes
        dirs[:] = [d for d in dirs if d not in {'.git', '__pycache__', 'node_modules', 'venv', '.venv'}]
        
        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if ext in extensiones_datos:
                archivos_encontrados.append(os.path.join(root, f))
    
    log(f"Encontrados {len(archivos_encontrados)} archivos de datos")
    
    # Analizar cada archivo
    for i, ruta in enumerate(archivos_encontrados):
        ext = os.path.splitext(ruta)[1].lower()
        nombre = os.path.basename(ruta)
        log(f"[{i+1}/{len(archivos_encontrados)}] Analizando: {nombre}")
        
        try:
            if ext == '.csv':
                resultado = analizar_csv(ruta, callback)
            elif ext in ('.xlsx', '.xls'):
                resultado = analizar_excel(ruta, callback)
            elif ext in ('.parquet', '.pq'):
                resultado = analizar_parquet(ruta, callback)
            else:
                continue
            
            resultados.append(resultado)
            
        except Exception as e:
            log(f"Error analizando {nombre}: {e}")
            resultados.append({'ruta': ruta, 'error': str(e)})
    
    # Generar reporte según formato
    
    # Crear carpeta de salida si no existe
    if not os.path.exists(carpeta_salida):
        try:
            os.makedirs(carpeta_salida)
        except OSError as e:
            log(f"Error creando carpeta de salida: {e}")
            return {'error': str(e)}
    
    ruta_salida = None
    
    try:
        if formato.lower() == 'txt':
            nombre_archivo = archivo_salida if archivo_salida.lower().endswith('.txt') else f"{archivo_salida.rsplit('.', 1)[0]}.txt"
            ruta_salida = os.path.join(carpeta_salida, nombre_archivo)
            _generar_reporte_txt(resultados, ruta_salida, log)
            
        elif formato.lower() == 'csv':
            nombre_archivo = archivo_salida if archivo_salida.lower().endswith('.csv') else f"{archivo_salida.rsplit('.', 1)[0]}.csv"
            ruta_salida = os.path.join(carpeta_salida, nombre_archivo)
            _generar_reporte_csv(resultados, ruta_salida, log)
            
        else: # Default PDF
            nombre_archivo = archivo_salida if archivo_salida.lower().endswith('.pdf') else f"{archivo_salida.rsplit('.', 1)[0]}.pdf"
            ruta_salida = os.path.join(carpeta_salida, nombre_archivo)
            _generar_reporte_pdf(resultados, ruta_salida, log)
            
    except Exception as e:
        log(f"Error generando reporte {formato.upper()}: {e}")
        ruta_salida = None
    
    return {
        'archivos_analizados': len(resultados),
        'ruta_reporte': ruta_salida,
        'resultados': resultados
    }


def _generar_reporte_pdf(resultados: List[Dict], ruta_salida: str, log: Callable):
    """Genera el reporte en formato PDF"""
    pdf = PDFDatos()
    pdf.add_page()
    
    for res in resultados:
        nombre = os.path.basename(res.get('ruta', 'N/A'))
        pdf.archivo_header(nombre, res.get('tipo', 'Desconocido'), res.get('ruta', 'N/A'))
        
        if 'error' in res:
            pdf.error(res['error'])
            pdf.ln(3)
            continue
        
        tamaño = res.get('tamaño_bytes', 0)
        pdf.info(f'Tamaño: {_formato_bytes(tamaño)}')
        
        # Para Excel con múltiples hojas
        if 'hojas' in res:
            for hoja in res['hojas']:
                pdf.hoja_header(hoja['nombre'])
                fila_header = hoja.get('fila_encabezado', 1)
                fila_datos = fila_header + 1
                pdf.info(f'  Encabezados en fila: {fila_header}')
                pdf.info(f'  Datos comienzan en fila: {fila_datos}')
                pdf.info(f'  Filas de datos: {hoja["total_filas"]:,}')
                pdf.info(f'  Columnas ({len(hoja["columnas"])}):')
                _escribir_columnas_pdf(pdf, hoja)
        else:
            fila_header = res.get('fila_encabezado', 1)
            fila_datos = fila_header + 1
            pdf.info(f'Encabezados en fila: {fila_header}')
            pdf.info(f'Datos comienzan en fila: {fila_datos}')
            pdf.info(f'Filas de datos: {res.get("total_filas", 0):,}')
            pdf.info(f'Columnas ({len(res.get("columnas", []))}):')
            _escribir_columnas_pdf(pdf, res)
        
        pdf.ln(3)
    
    # Resumen final
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 8, f'Total archivos analizados: {len(resultados)}', 0, 1)
    
    pdf.output(ruta_salida)
    
    if os.path.exists(ruta_salida):
            log(f"Reporte PDF guardado en: {ruta_salida}")
    else:
            log(f"Error CRÍTICO: El reporte no aparece en {ruta_salida}")


def _generar_reporte_txt(resultados: List[Dict], ruta_salida: str, log: Callable):
    """Genera el reporte en formato TXT"""
    with open(ruta_salida, 'w', encoding='utf-8') as f:
        f.write("MAPA DE ARCHIVOS DE DATOS\n")
        f.write("=========================\n\n")
        
        for res in resultados:
            nombre = os.path.basename(res.get('ruta', 'N/A'))
            tipo = res.get('tipo', 'Desconocido')
            ruta = res.get('ruta', 'N/A')
            
            f.write(f"ARCHIVO: {nombre}\n")
            f.write(f"Tipo: {tipo}\n")
            f.write(f"Ruta: {ruta}\n")
            
            if 'error' in res:
                f.write(f"ERROR: {res['error']}\n\n")
                continue
            
            tamaño = res.get('tamaño_bytes', 0)
            f.write(f"Tamaño: {_formato_bytes(tamaño)}\n")
            
            if 'hojas' in res:
                for hoja in res['hojas']:
                    f.write(f"\n  HOJA: {hoja['nombre']}\n")
                    f.write(f"  Encabezados en fila: {hoja.get('fila_encabezado', 1)}\n")
                    f.write(f"  Filas de datos: {hoja.get('total_filas', 0):,}\n")
                    f.write(f"  Columnas ({len(hoja.get('columnas', []))}):\n")
                    _escribir_columnas(f, hoja, indent="    ")
            else:
                f.write(f"Encabezados en fila: {res.get('fila_encabezado', 1)}\n")
                f.write(f"Filas de datos: {res.get('total_filas', 0):,}\n")
                f.write(f"Columnas ({len(res.get('columnas', []))}):\n")
                _escribir_columnas(f, res)
            
            f.write("\n" + "-"*50 + "\n\n")
            
        f.write(f"Total archivos analizados: {len(resultados)}\n")
        
    if os.path.exists(ruta_salida):
        log(f"Reporte TXT guardado en: {ruta_salida}")
    else:
        log(f"Error CRÍTICO: El reporte no aparece en {ruta_salida}")


def _generar_reporte_csv(resultados: List[Dict], ruta_salida: str, log: Callable):
    """Genera el reporte en formato CSV (Flat format)"""
    import csv
    
    # Definir columnas del reporte CSV
    fieldnames = [
        'archivo_nombre', 'archivo_ruta', 'archivo_tipo', 'archivo_tamaño', 
        'hoja_nombre', 'total_filas', 'total_columnas', 
        'columna_nombre', 'columna_tipo_detectado', 'ejemplo_valores', 'error'
    ]
    
    with open(ruta_salida, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        
        for res in resultados:
            base_info = {
                'archivo_nombre': os.path.basename(res.get('ruta', 'N/A')),
                'archivo_ruta': res.get('ruta', 'N/A'),
                'archivo_tipo': res.get('tipo', 'Desconocido'),
                'archivo_tamaño': _formato_bytes(res.get('tamaño_bytes', 0)),
                'error': res.get('error', '')
            }
            
            if 'error' in res:
                writer.writerow(base_info)
                continue
            
            # Helper para escribir columnas
            def escribir_cols(columns_data, hoja_nom=''):
                columnas = columns_data.get('columnas', [])
                tipos = columns_data.get('tipos_detectados', {})
                muestras = columns_data.get('muestra_valores', {})
                
                if not columnas:
                    row = base_info.copy()
                    row.update({
                        'hoja_nombre': hoja_nom,
                        'total_filas': columns_data.get('total_filas', 0),
                        'total_columnas': 0
                    })
                    writer.writerow(row)
                    return

                for col in columnas:
                    row = base_info.copy()
                    muestra = muestras.get(col, [])
                    muestra_str = "; ".join(muestra[:5]) # Separador interno ;
                    
                    row.update({
                        'hoja_nombre': hoja_nom,
                        'total_filas': columns_data.get('total_filas', 0),
                        'total_columnas': len(columnas),
                        'columna_nombre': col,
                        'columna_tipo_detectado': tipos.get(col, 'texto'),
                        'ejemplo_valores': muestra_str
                    })
                    writer.writerow(row)
            
            if 'hojas' in res:
                for hoja in res['hojas']:
                    escribir_cols(hoja, hoja['nombre'])
            else:
                escribir_cols(res)

    if os.path.exists(ruta_salida):
        log(f"Reporte CSV guardado en: {ruta_salida}")
    else:
        log(f"Error CRÍTICO: El reporte no aparece en {ruta_salida}")


def _escribir_columnas_pdf(pdf, datos: Dict):
    """Escribe información de columnas al PDF"""
    columnas = datos.get('columnas', [])
    tipos = datos.get('tipos_detectados', {})
    muestras = datos.get('muestra_valores', {})
    
    for col in columnas:
        tipo = tipos.get(col, 'texto')
        muestra = muestras.get(col, [])
        muestra_str = ", ".join(muestra[:3]) if muestra else "(vacio)"
        if len(muestra_str) > 50:
            muestra_str = muestra_str[:47] + "..."
        pdf.columna(col, tipo, muestra_str)


def _escribir_columnas(f, datos: Dict, indent: str = "  "):
    """Escribe información de columnas al archivo"""
    columnas = datos.get('columnas', [])
    tipos = datos.get('tipos_detectados', {})
    muestras = datos.get('muestra_valores', {})
    
    for col in columnas:
        tipo = tipos.get(col, 'texto')
        muestra = muestras.get(col, [])
        muestra_str = ", ".join(muestra[:3]) if muestra else "(vacío)"
        if len(muestra_str) > 60:
            muestra_str = muestra_str[:57] + "..."
        f.write(f"{indent}- {col} [{tipo}]: {muestra_str}\n")


def _formato_bytes(bytes_num: int) -> str:
    """Formatea bytes a unidad legible"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if bytes_num < 1024:
            return f"{bytes_num:.1f} {unit}"
        bytes_num /= 1024
    return f"{bytes_num:.1f} TB"


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        directorio = sys.argv[1]
    else:
        directorio = "."
    
    generar_reporte_datos(directorio)
